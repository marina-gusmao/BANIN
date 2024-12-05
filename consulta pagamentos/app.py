#verificar se valores estão pagos ou não e, caso estejam, verificar a forma de pagamento
#usar um site para verificar se está pago ou não, a data, e a forma de pagamento
#com base em uma planilha, pegar o cpf para verificar o pagamento
#se estiver pago, preencher a planilha de fechamento como ok
#se não, informar que continua pendente
#1. definir quais são os passos manuais dessa tarefa para transformá-las em código manualmente

# ----- 1 entrar na planilha e extrair o cpf do cliente
# ----- 2 entrar no site e usar cpf da planilha para pesquisar o status de pagamento
# ----- 3 verificar se está em dia ou atrasado
# -----	4 se estiver em dia, pegar a data do pagamento e o método de pagamento
# ----- 5 caso contrário, se estiver atrasado, colocar o status como pendente
# ----- 6 inserir essas novas informações (nome, valor, cpf, vencimento, status e, caso em dia, data pagamento e método pagamento) em uma nova planilha
# ----- 7 repetir até chegar no último cliente

import openpyxl
from selenium import webdriver #permite abrir o navegador
from selenium.webdriver.common.by import By #permite interagir com a página
from time import sleep #permite pausar o programa por tempo determinado

# ----- 1 entrar na planilha e extrair o cpf do cliente

planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app') # ----- 2 entrar no site e usar cpf da planilha para pesquisar o status de pagamento


 #função q permite ler cada linha, podendo especificar onde começar a leitura
 #a linha pra começar a leitura, nesse caso, é a 2, pulando o cabeçalho
for linha in pagina_clientes.iter_rows(min_row=2,values_only=True):
    nome, valor, cpf, vencimento = linha
    sleep(5) #pra garantir que o site carregou
    campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']") #o xpath é onde está o elemento a ser consultado quando inspeciona a página, lá mostra o @atributo='valor'
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    # ----- 3 ferificar se está em dia ou atrasado
    botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']") #encontra o botão de busca
    sleep(1)
    botao_pesquisar.click() #clica no botão de busca
    sleep(4)
    # ----- 3 verificar se está em dia ou atrasado
    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    if status.text == 'em dia':#mostra o texto extraído
        # -----	4 se estiver em dia, pegar a data do pagamento e o método de pagamento
        data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
        data_pagamento_limpo = data_pagamento.text.split()[-1]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[-1]
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])
        planilha_fechamento.save('planilha fechamento.xlsx')
    else:
        # ----- 5 caso contrário, se estiver atrasado, colocar o status como pendente
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        planilha_fechamento.save('planilha fechamento.xlsx')